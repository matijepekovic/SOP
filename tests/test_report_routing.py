from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from sop_reporter.config import (
    ConfigurationError,
    load_report_definitions,
)
from sop_reporter.email_client import DownloadedAttachment, EmailRecord
from sop_reporter.exceptions import ExtractionError
from sop_reporter.extractor import ExtractedData, ExtractedPartition
from sop_reporter.pipeline import JobRunner, RunStatus
from sop_reporter.state_store import ProcessedStateStore
from tests.test_pipeline import FakeBuilder, FakePrinter, make_job


PROJECT_ROOT = Path(__file__).resolve().parent.parent


class RecordingExtractor:
    """Extractor that records which files it was handed."""

    def __init__(self, label: str = "", rows: int = 1) -> None:
        self.label = label
        self.rows = rows
        self.seen: list[str] = []

    def extract_partitions(self, paths):
        source_files = tuple(paths)
        self.seen.extend(path.name for path in source_files)
        data = ExtractedData(
            headers=("Job Number",),
            rows=tuple({"Job Number": f"J{i}"} for i in range(self.rows)),
            number_formats={},
            widths={"Job Number": 14},
            source_files=source_files,
        )
        return (ExtractedPartition(self.label, data),)


class FailingExtractor(RecordingExtractor):
    def extract_partitions(self, paths):
        self.seen.extend(path.name for path in paths)
        raise ExtractionError("missing required column(s): Sub Status")


class MultiPartitionExtractor(RecordingExtractor):
    def extract_partitions(self, paths):
        source_files = tuple(paths)
        self.seen.extend(path.name for path in source_files)

        def partition(label: str) -> ExtractedPartition:
            return ExtractedPartition(
                label,
                ExtractedData(
                    headers=("Job Number",),
                    rows=({"Job Number": label},),
                    number_formats={},
                    widths={"Job Number": 14},
                    source_files=source_files,
                ),
            )

        return (
            partition("Item Notification"),
            partition("Install Issue"),
            partition("On Hold"),
        )


class MultiAttachmentEmailClient:
    def __init__(self, record: EmailRecord) -> None:
        self.record = record

    def fetch_messages(self, _download_dir, handled_message_ids=()):
        return [] if self.record.message_id in handled_message_ids else [self.record]


def _record_with(root: Path, filenames: list[str]) -> EmailRecord:
    attachments = []
    for filename in filenames:
        path = root / filename
        path.write_bytes(b"input")
        attachments.append(
            DownloadedAttachment(
                filename,
                path,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                5,
            )
        )
    return EmailRecord(
        sequence_id="1",
        message_id="<one@example>",
        subject="Morning SOP",
        sender="reports@example.com",
        sent_date="Thu, 13 Aug 2026 07:00:00 -0700",
        attachments=tuple(attachments),
    )


def _runner(root: Path, record: EmailRecord, jobs, printer) -> tuple[JobRunner, ProcessedStateStore]:
    store = ProcessedStateStore(root / "state.json")
    runner = JobRunner(
        email_client=MultiAttachmentEmailClient(record),
        state_store=store,
        jobs=jobs,
        printer=printer,
        downloads_dir=root / "downloads",
        reports_dir=root / "reports",
        report_filename="Report_%Y%m%d_%H%M%S.xlsx",
    )
    return runner, store


class RoutingTests(unittest.TestCase):
    def test_each_attachment_goes_only_to_its_own_report(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            record = _record_with(
                root,
                [
                    "GM SOP-2026-08-13.xlsx",
                    "Morning SOP ALL-2026-08-13.xlsx",
                    "need to collect Company-2026-08-13.xlsx",
                ],
            )
            gm = RecordingExtractor()
            morning = RecordingExtractor()
            collect = RecordingExtractor()
            jobs = [
                make_job("GM SOP", gm, patterns=("GM SOP*.xlsx",)),
                make_job("Morning SOP ALL", morning, patterns=("Morning SOP ALL*.xlsx",)),
                make_job(
                    "Need To Collect Company",
                    collect,
                    patterns=("need to collect Company*.xlsx",),
                ),
            ]
            printer = FakePrinter()
            runner, _ = _runner(root, record, jobs, printer)

            result = runner.run_once()

            self.assertEqual(result.status, RunStatus.SUCCESS)
            self.assertEqual(gm.seen, ["GM SOP-2026-08-13.xlsx"])
            self.assertEqual(morning.seen, ["Morning SOP ALL-2026-08-13.xlsx"])
            self.assertEqual(collect.seen, ["need to collect Company-2026-08-13.xlsx"])
            self.assertEqual(len(printer.printed), 3)

    def test_matching_is_case_insensitive(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            record = _record_with(root, ["gm sop-LOWER.xlsx"])
            gm = RecordingExtractor()
            runner, _ = _runner(
                root,
                record,
                [make_job("GM SOP", gm, patterns=("GM SOP*.xlsx",))],
                FakePrinter(),
            )

            runner.run_once()

            self.assertEqual(gm.seen, ["gm sop-LOWER.xlsx"])

    def test_unmatched_attachment_is_skipped_without_failing_the_run(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            record = _record_with(
                root, ["GM SOP-1.xlsx", "Some Unrelated Spreadsheet.xlsx"]
            )
            gm = RecordingExtractor()
            printer = FakePrinter()
            runner, _ = _runner(
                root,
                record,
                [make_job("GM SOP", gm, patterns=("GM SOP*.xlsx",))],
                printer,
            )

            result = runner.run_once()

            self.assertEqual(result.status, RunStatus.SUCCESS)
            self.assertEqual(gm.seen, ["GM SOP-1.xlsx"])
            self.assertEqual(len(printer.printed), 1)

    def test_disabled_report_never_receives_attachments(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            record = _record_with(root, ["Morning SOP ALL-1.xlsx", "GM SOP-1.xlsx"])
            morning = RecordingExtractor()
            gm = RecordingExtractor()
            runner, _ = _runner(
                root,
                record,
                [
                    make_job(
                        "Morning SOP ALL",
                        morning,
                        patterns=("Morning SOP ALL*.xlsx",),
                        enabled=False,
                    ),
                    make_job("GM SOP", gm, patterns=("GM SOP*.xlsx",)),
                ],
                FakePrinter(),
            )

            runner.run_once()

            self.assertEqual(morning.seen, [])
            self.assertEqual(gm.seen, ["GM SOP-1.xlsx"])

    def test_one_failing_report_still_lets_the_others_print(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            record = _record_with(root, ["GM SOP-1.xlsx", "Morning SOP ALL-1.xlsx"])
            gm = RecordingExtractor()
            morning = FailingExtractor()
            printer = FakePrinter()
            runner, store = _runner(
                root,
                record,
                [
                    make_job("GM SOP", gm, patterns=("GM SOP*.xlsx",)),
                    make_job(
                        "Morning SOP ALL", morning, patterns=("Morning SOP ALL*.xlsx",)
                    ),
                ],
                printer,
            )

            with self.assertLogs("sop_reporter.pipeline", level="ERROR"):
                result = runner.run_once()

            # The healthy report still printed...
            self.assertEqual(result.status, RunStatus.SUCCESS)
            self.assertEqual(len(printer.printed), 1)
            # ...and the failure is named rather than swallowed.
            self.assertIn("Morning SOP ALL", result.message)
            self.assertIn("Sub Status", result.message)
            self.assertTrue(store.is_handled("<one@example>"))

    def test_every_report_failing_fails_the_run(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            record = _record_with(root, ["GM SOP-1.xlsx"])
            printer = FakePrinter()
            runner, _ = _runner(
                root,
                record,
                [make_job("GM SOP", FailingExtractor(), patterns=("GM SOP*.xlsx",))],
                printer,
            )

            with self.assertLogs("sop_reporter.pipeline", level="ERROR"):
                result = runner.run_once()

            self.assertEqual(result.status, RunStatus.FAILED)
            self.assertEqual(printer.printed, [])
            self.assertIn("GM SOP", result.message)

    def test_sub_status_partitions_each_become_their_own_printout(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            record = _record_with(root, ["GM SOP-1.xlsx"])
            printer = FakePrinter()
            runner, _ = _runner(
                root,
                record,
                [
                    make_job(
                        "GM SOP",
                        MultiPartitionExtractor(),
                        patterns=("GM SOP*.xlsx",),
                    )
                ],
                printer,
            )

            result = runner.run_once()

            self.assertEqual(len(result.report_paths), 3)
            self.assertEqual(len(printer.printed), 3)
            names = [path.name for path in result.report_paths]
            self.assertTrue(any("Item_Notification" in name for name in names))
            self.assertTrue(any("Install_Issue" in name for name in names))
            self.assertTrue(any("On_Hold" in name for name in names))

    def test_per_report_filename_keeps_outputs_from_colliding(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            record = _record_with(root, ["GM SOP-1.xlsx", "Morning SOP ALL-1.xlsx"])
            runner, _ = _runner(
                root,
                record,
                [
                    make_job(
                        "GM SOP",
                        RecordingExtractor(),
                        patterns=("GM SOP*.xlsx",),
                        report_filename="GM_SOP_%Y%m%d.xlsx",
                    ),
                    make_job(
                        "Morning SOP ALL",
                        RecordingExtractor(),
                        patterns=("Morning SOP ALL*.xlsx",),
                        report_filename="Morning_SOP_ALL_%Y%m%d.xlsx",
                    ),
                ],
                FakePrinter(),
            )

            result = runner.run_once()

            names = [path.name for path in result.report_paths]
            self.assertEqual(len(set(names)), 2)
            self.assertTrue(any(name.startswith("GM_SOP_") for name in names))
            self.assertTrue(any(name.startswith("Morning_SOP_ALL_") for name in names))


class DefinitionLoadingTests(unittest.TestCase):
    def test_shipped_rules_load_with_gm_sop_enabled(self) -> None:
        definitions = load_report_definitions(PROJECT_ROOT / "config")
        by_name = {d.name: d for d in definitions}

        for name in ("GM SOP", "Morning SOP ALL", "Need To Collect Company"):
            self.assertIn(name, by_name)
            self.assertTrue(by_name[name].match.enabled)
        # GM SOP and Morning SOP ALL print one document per Sub Status.
        self.assertTrue(by_name["GM SOP"].extraction.split.enabled)
        self.assertTrue(by_name["Morning SOP ALL"].extraction.split.enabled)
        # Need To Collect is already filtered to one Sub Status in Salesforce
        # and has no Sub Status column, so it prints as a single document.
        self.assertFalse(by_name["Need To Collect Company"].extraction.split.enabled)
        # Every report writes to its own filename so outputs cannot collide.
        filenames = [d.report_filename for d in by_name.values()]
        self.assertEqual(len(filenames), len(set(filenames)))

    def test_the_two_live_reports_claim_different_files(self) -> None:
        by_name = {d.name: d for d in load_report_definitions(PROJECT_ROOT / "config")}
        gm = by_name["GM SOP"].match
        morning = by_name["Morning SOP ALL"].match
        self.assertTrue(gm.matches("GM SOP-2026-08-13.xlsx"))
        self.assertFalse(gm.matches("Morning SOP ALL-2026-08-13.xlsx"))
        self.assertTrue(morning.matches("Morning SOP ALL-2026-08-13.xlsx"))
        self.assertFalse(morning.matches("GM SOP-2026-08-13.xlsx"))
        collect = by_name["Need To Collect Company"].match
        self.assertTrue(collect.matches("need to collect Company-2026-08-13.xlsx"))
        self.assertFalse(collect.matches("GM SOP-2026-08-13.xlsx"))

    def test_missing_rules_directory_falls_back_to_the_legacy_file(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            config_dir = Path(directory)
            legacy = config_dir / "extraction_rules.yaml"
            legacy.write_bytes(
                (PROJECT_ROOT / "config" / "extraction_rules.default.yaml").read_bytes()
            )

            definitions = load_report_definitions(config_dir)

            self.assertEqual(len(definitions), 1)
            self.assertEqual(definitions[0].name, "extraction_rules")
            self.assertTrue(definitions[0].match.enabled)

    def test_no_rules_at_all_is_a_clear_configuration_error(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaises(ConfigurationError) as caught:
                load_report_definitions(Path(directory))
            self.assertIn("No report rules found", str(caught.exception))

    def test_a_broken_rule_file_names_itself(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            config_dir = Path(directory)
            rules = config_dir / "rules"
            rules.mkdir()
            (rules / "broken.yaml").write_text("version: 1\ninput: {}\n", encoding="utf-8")

            with self.assertRaises(ConfigurationError) as caught:
                load_report_definitions(config_dir)
            self.assertIn("broken.yaml", str(caught.exception))


if __name__ == "__main__":
    unittest.main()


class ContentMatchingTests(unittest.TestCase):
    """Filenames from the sending system are not stable, so a report must be
    identifiable by the columns its workbook contains."""

    def _workbook(self, root: Path, name: str, *, lead_rows: int) -> Path:
        from openpyxl import Workbook, load_workbook

        source = load_workbook(
            PROJECT_ROOT / "tests" / "fixtures" / "salesforce_olympia_sample.xlsx"
        )
        rows = list(source.active.iter_rows(values_only=True))
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "GM SOP"
        for index in range(lead_rows):
            # Title and filter banners of the kind report exports put above
            # the real header row.
            sheet.append(["GM SOP Report"] if index == 1 else [])
        for row in rows:
            sheet.append(list(row))
        path = root / name
        workbook.save(path)
        return path

    def _gm_job(self, *, patterns=("GM SOP*.xlsx",)):
        from sop_reporter.extractor import ExtractionEngine

        definition = {
            d.name: d for d in load_report_definitions(PROJECT_ROOT / "config")
        }["GM SOP"]
        return make_job(
            "GM SOP",
            ExtractionEngine(definition.extraction),
            patterns=patterns,
        )

    def test_header_below_blank_and_title_rows_is_found(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            path = self._workbook(root, "GM SOP-1.xlsx", lead_rows=4)
            job = self._gm_job()

            partitions = job.extractor.extract_partitions([path])

            self.assertEqual(
                [p.label for p in partitions],
                ["Item Notification", "Install Issue", "On Hold"],
            )

    def test_unrecognizable_filename_still_routes_by_columns(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            # Nothing about this name matches the configured patterns.
            path = self._workbook(root, "d080ccd547be_01_export.xlsx", lead_rows=4)
            record = EmailRecord(
                sequence_id="1",
                message_id="<one@example>",
                subject="SOP",
                sender="reports@example.com",
                sent_date="Thu, 13 Aug 2026 07:00:00 -0700",
                attachments=(
                    DownloadedAttachment(
                        "d080ccd547be_01_export.xlsx",
                        path,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        5,
                    ),
                ),
            )
            printer = FakePrinter()
            runner, _ = _runner(root, record, [self._gm_job()], printer)

            result = runner.run_once()

            self.assertEqual(result.status, RunStatus.SUCCESS)
            self.assertEqual(len(printer.printed), 3)

    def test_a_workbook_without_the_columns_is_not_claimed(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = Workbook()
            workbook.active.append(["Something", "Entirely", "Different"])
            path = root / "mystery.xlsx"
            workbook.save(path)

            self.assertFalse(self._gm_job().extractor.can_handle(path))
