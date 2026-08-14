from __future__ import annotations

import logging
import math
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import date, datetime
from functools import cmp_to_key
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook

from sop_reporter.config import (
    AggregationRule,
    ColumnRule,
    ExtractionConfig,
    FilterRule,
)
from sop_reporter.exceptions import ExtractionError


LOGGER = logging.getLogger(__name__)
NUMBER_CLEANUP = re.compile(r"[^0-9.\-+]" )


@dataclass(frozen=True)
class ExtractedData:
    headers: tuple[str, ...]
    rows: tuple[dict[str, Any], ...]
    number_formats: Mapping[str, str]
    widths: Mapping[str, float]
    source_files: tuple[Path, ...]


@dataclass(frozen=True)
class ExtractedPartition:
    label: str
    data: ExtractedData


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


# Report exports mark sorted columns by appending an arrow to the header, so
# "Market" arrives as "Market  ↑". Which columns carry one changes with the
# sort order, so the arrows are stripped rather than written into the rules.
SORT_INDICATORS = "↑↓▲▼⬆⬇▴▾⇅↕"
_HEADER_WHITESPACE = re.compile(r"\s+")


def normalize_header(value: Any) -> str:
    """Fold a header cell to the form column rules are written in."""
    text = str(value).replace(" ", " ")
    text = text.strip().strip(SORT_INDICATORS + " \t")
    return _HEADER_WHITESPACE.sub(" ", text).strip()


def header_key(value: Any) -> str:
    return normalize_header(value).casefold()


def _normalized_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_summary_label(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    normalized = value.strip().casefold()
    return (
        normalized.startswith("subtotal")
        or normalized.startswith("total")
        or normalized.startswith("grand total")
    )


def _parse_number(value: Any) -> float | int | None:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        raise ValueError("boolean is not a number")
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            raise ValueError("number is not finite")
        return value
    text = str(value).strip()
    negative_parentheses = text.startswith("(") and text.endswith(")")
    cleaned = NUMBER_CLEANUP.sub("", text)
    if cleaned in {"", "+", "-", "."}:
        raise ValueError(f"not a number: {value!r}")
    parsed = float(cleaned)
    if negative_parentheses:
        parsed = -abs(parsed)
    return parsed


def _parse_date(value: Any, include_time: bool) -> date | datetime | None:
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value if include_time else value.date()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()) if include_time else value
    text = str(value).strip()
    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d-%b-%Y",
    )
    for format_string in formats:
        try:
            parsed = datetime.strptime(text, format_string)
            return parsed if include_time else parsed.date()
        except ValueError:
            continue
    raise ValueError(f"not a supported date: {value!r}")


def _convert_value(value: Any, rule: ColumnRule) -> Any:
    if _is_blank(value):
        value = rule.default
    if _is_blank(value):
        return None
    kind = rule.value_type
    try:
        if kind == "text":
            return _normalized_text(value)
        if kind == "integer":
            parsed = _parse_number(value)
            if parsed is None:
                return None
            if float(parsed).is_integer():
                return int(parsed)
            raise ValueError(f"not a whole number: {value!r}")
        if kind in {"number", "currency"}:
            return _parse_number(value)
        if kind == "percent":
            if isinstance(value, str) and "%" in value:
                parsed = _parse_number(value)
                return None if parsed is None else float(parsed) / 100.0
            return _parse_number(value)
        if kind == "date":
            return _parse_date(value, include_time=False)
        if kind == "datetime":
            return _parse_date(value, include_time=True)
        if kind == "boolean":
            if isinstance(value, bool):
                return value
            normalized = str(value).strip().casefold()
            if normalized in {"true", "yes", "y", "1"}:
                return True
            if normalized in {"false", "no", "n", "0"}:
                return False
            raise ValueError(f"not a boolean: {value!r}")
    except (TypeError, ValueError) as exc:
        raise ExtractionError(
            f"Cannot convert value {value!r} in '{rule.source}' to {kind}: {exc}"
        ) from exc
    raise ExtractionError(f"Unsupported conversion type: {kind}")


def _lookup_case_insensitive(row: Mapping[str, Any], column: str) -> Any:
    if column in row:
        return row[column]
    folded = column.casefold()
    matches = [value for key, value in row.items() if key.casefold() == folded]
    if not matches:
        raise ExtractionError(f"Filter references unknown column: {column}")
    return matches[0]


def _coerce_expected(actual: Any, expected: Any) -> Any:
    if _is_blank(expected):
        return expected
    if isinstance(actual, bool):
        if isinstance(expected, bool):
            return expected
        normalized = str(expected).strip().casefold()
        if normalized in {"true", "yes", "y", "1"}:
            return True
        if normalized in {"false", "no", "n", "0"}:
            return False
    if isinstance(actual, (int, float)) and not isinstance(actual, bool):
        return _parse_number(expected)
    if isinstance(actual, datetime):
        return _parse_date(expected, include_time=True)
    if isinstance(actual, date):
        return _parse_date(expected, include_time=False)
    return expected


def _equal(actual: Any, expected: Any, case_sensitive: bool) -> bool:
    if _is_blank(actual) or _is_blank(expected):
        return _is_blank(actual) and _is_blank(expected)
    coerced = _coerce_expected(actual, expected)
    if isinstance(actual, str) or isinstance(coerced, str):
        left = str(actual).strip()
        right = str(coerced).strip()
        return left == right if case_sensitive else left.casefold() == right.casefold()
    return actual == coerced


def _ordered_compare(actual: Any, expected: Any) -> int:
    coerced = _coerce_expected(actual, expected)
    if isinstance(actual, str) or isinstance(coerced, str):
        left = str(actual).casefold()
        right = str(coerced).casefold()
    else:
        left = actual
        right = coerced
    if left < right:
        return -1
    if left > right:
        return 1
    return 0


def _matches_filter(row: Mapping[str, Any], rule: FilterRule) -> bool:
    actual = _lookup_case_insensitive(row, rule.column)
    operator = rule.operator
    if operator == "is_blank":
        return _is_blank(actual)
    if operator == "not_blank":
        return not _is_blank(actual)
    if operator == "equals":
        return _equal(actual, rule.value, rule.case_sensitive)
    if operator == "not_equals":
        return not _equal(actual, rule.value, rule.case_sensitive)
    if operator in {"contains", "not_contains"}:
        if _is_blank(actual):
            contained = False
        else:
            left = str(actual)
            right = str(rule.value)
            if not rule.case_sensitive:
                left, right = left.casefold(), right.casefold()
            contained = right in left
        return contained if operator == "contains" else not contained
    if operator in {"in", "not_in"}:
        contained = any(
            _equal(actual, expected, rule.case_sensitive) for expected in rule.values
        )
        return contained if operator == "in" else not contained
    if _is_blank(actual):
        return False
    try:
        if operator == "range":
            lower_ok = (
                True
                if rule.minimum is None
                else _ordered_compare(actual, rule.minimum) >= 0
            )
            upper_ok = (
                True
                if rule.maximum is None
                else _ordered_compare(actual, rule.maximum) <= 0
            )
            return lower_ok and upper_ok
        comparison = _ordered_compare(actual, rule.value)
        return {
            "gt": comparison > 0,
            "gte": comparison >= 0,
            "lt": comparison < 0,
            "lte": comparison <= 0,
        }[operator]
    except (TypeError, ValueError) as exc:
        raise ExtractionError(
            f"Cannot apply filter {operator} to column {rule.column}: {exc}"
        ) from exc


class ExtractionEngine:
    def __init__(self, config: ExtractionConfig) -> None:
        self.config = config

    def extract_file(self, path: Path) -> ExtractedData:
        return self.extract_files([path])

    def extract_files(self, paths: Iterable[Path]) -> ExtractedData:
        source_files = tuple(Path(path) for path in paths)
        mapped_rows = self._read_all_mapped_rows(source_files)
        return self._finalize(mapped_rows, source_files)

    def extract_partitions(self, paths: Iterable[Path]) -> tuple[ExtractedPartition, ...]:
        source_files = tuple(Path(path) for path in paths)
        mapped_rows = self._read_all_mapped_rows(source_files)
        split = self.config.split
        if not split.enabled:
            return (
                ExtractedPartition(
                    label="",
                    data=self._finalize(mapped_rows, source_files),
                ),
            )

        assert split.by is not None
        grouped: OrderedDict[str, tuple[str, list[dict[str, Any]]]] = OrderedDict()
        for row in mapped_rows:
            value = row.get(split.by)
            if _is_blank(value):
                if not split.include_blank:
                    continue
                label = split.blank_label
                key = "__blank__"
            else:
                label = _normalized_text(value)
                key = label.casefold()
            if key not in grouped:
                grouped[key] = (label, [])
            grouped[key][1].append(row)

        return tuple(
            ExtractedPartition(
                label=label,
                data=self._finalize(rows, source_files),
            )
            for label, rows in grouped.values()
        )

    def _read_all_mapped_rows(
        self, source_files: Sequence[Path]
    ) -> list[dict[str, Any]]:
        mapped_rows: list[dict[str, Any]] = []
        for path in source_files:
            mapped_rows.extend(self._read_mapped_rows(path))
        return mapped_rows

    HEADER_SCAN_ROWS = 25

    def can_handle(self, path: Path) -> bool:
        """Whether this report's required columns exist in ``path``.

        Lets a workbook be identified by what is inside it rather than by its
        filename, which matters because the sending system's filenames are not
        stable.
        """
        if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
            return False
        required = {
            header_key(rule.source) for rule in self.config.columns if rule.required
        }
        if not required:
            return False
        try:
            workbook = load_workbook(
                path, read_only=True, data_only=True, keep_links=False
            )
        except Exception:
            return False
        try:
            worksheet = self._select_worksheet(workbook, path)
            _row, values = self._locate_header_row(worksheet, path)
        except ExtractionError:
            return False
        finally:
            workbook.close()
        present = {header_key(value) for value in values if not _is_blank(value)}
        return required.issubset(present)

    def _locate_header_row(self, worksheet, path: Path) -> tuple[int, tuple]:
        """Return the header row number and its cell values.

        With an explicit ``input.header_row`` this just reads that row. With
        ``header_row: auto`` it scans the top of the sheet and picks the row
        matching the most configured source columns, which is what exports
        that begin with title or filter rows need.
        """
        configured = self.config.input.header_row
        if configured:
            values = next(
                worksheet.iter_rows(
                    min_row=configured, max_row=configured, values_only=True
                ),
                (),
            )
            return configured, values

        wanted = {header_key(rule.source) for rule in self.config.columns}
        required = {
            header_key(rule.source) for rule in self.config.columns if rule.required
        }
        best_row = 0
        best_values: tuple = ()
        best_score = 0
        scanned: list[str] = []

        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=1, max_row=self.HEADER_SCAN_ROWS, values_only=True
            ),
            start=1,
        ):
            labels = {
                header_key(value) for value in values if not _is_blank(value)
            }
            if labels:
                preview = ", ".join(
                    str(v).strip() for v in values if not _is_blank(v)
                )
                scanned.append(f"row {row_number}: {preview[:160]}")
            score = len(labels & wanted)
            if score > best_score:
                best_score, best_row, best_values = score, row_number, values

        # Require every mandatory column, so a stray row that happens to echo
        # one or two column names is never mistaken for the header.
        if best_row and required:
            found = {header_key(v) for v in best_values if not _is_blank(v)}
            if not required.issubset(found):
                best_row = 0

        if not best_row:
            detail = "; ".join(scanned) if scanned else "(all scanned rows are blank)"
            raise ExtractionError(
                f"Could not find the header row in {path.name} sheet "
                f"{worksheet.title!r}. Looked for columns "
                f"{', '.join(sorted(rule.source for rule in self.config.columns if rule.required))} "
                f"in the first {self.HEADER_SCAN_ROWS} rows. "
                f"Rows seen: {detail}. "
                "Set input.header_row to the correct row number, or correct the "
                "column sources, in this report's rules file."
            )

        LOGGER.info(
            "%s: detected header on row %d of sheet %r",
            path.name,
            best_row,
            worksheet.title,
        )
        return best_row, best_values

    def _read_mapped_rows(self, path: Path) -> list[dict[str, Any]]:
        if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
            raise ExtractionError(
                f"Unsupported attachment format {path.suffix!r}; save legacy .xls files as .xlsx"
            )
        try:
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as exc:
            raise ExtractionError(f"Cannot open workbook {path.name}: {exc}") from exc

        try:
            worksheet = self._select_worksheet(workbook, path)
            header_row, header_values = self._locate_header_row(worksheet, path)
            data_start_row = self.config.input.data_start_row or header_row + 1
            headers: dict[str, tuple[str, int]] = {}
            for index, value in enumerate(header_values):
                if _is_blank(value):
                    continue
                name = normalize_header(value)
                folded = name.casefold()
                if folded in headers:
                    raise ExtractionError(
                        f"Duplicate header '{name}' in {path.name} row {header_row}"
                    )
                headers[folded] = (name, index)

            found = [name for name, _index in headers.values()]
            LOGGER.info(
                "%s sheet %r row %d headers: %s",
                path.name,
                worksheet.title,
                header_row,
                ", ".join(found) if found else "(none)",
            )

            missing = [
                rule.source
                for rule in self.config.columns
                if rule.required and header_key(rule.source) not in headers
            ]
            if missing:
                # Report what the workbook actually contains. Without this the
                # error only says what was wanted, which is a dead end when the
                # rules were written against a different report or the header
                # sits on another row.
                detail = ", ".join(found) if found else "(no non-blank cells)"
                raise ExtractionError(
                    f"Workbook {path.name} is missing required column(s): "
                    f"{', '.join(missing)}. "
                    f"Row {header_row} of sheet "
                    f"{worksheet.title!r} actually contains: {detail}. "
                    f"Sheets in this workbook: {', '.join(workbook.sheetnames)}. "
                    "Adjust input.header_row, input.sheet_name, or the column "
                    "sources in extraction_rules.yaml, or narrow "
                    "email.attachments.filename_patterns so only the intended "
                    "report is processed."
                )

            known_filter_columns = set(headers)
            known_filter_columns.update(
                header_key(rule.target) for rule in self.config.columns
            )
            unknown_filters = [
                rule.column
                for rule in self.config.filters.rules
                if rule.column.casefold() not in known_filter_columns
            ]
            if unknown_filters:
                raise ExtractionError(
                    f"Filter references unknown column(s) in {path.name}: {', '.join(unknown_filters)}"
                )

            output_rows: list[dict[str, Any]] = []
            fill_down_values: dict[str, Any] = {}
            for row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=data_start_row,
                    values_only=True,
                ),
                start=data_start_row,
            ):
                if all(_is_blank(value) for value in values):
                    if self.config.input.stop_at_first_blank_row:
                        break
                    if self.config.input.skip_blank_rows:
                        continue

                raw_row = {
                    original: values[index] if index < len(values) else None
                    for original, index in headers.values()
                }
                mapped_row: dict[str, Any] = {}
                try:
                    for rule in self.config.columns:
                        header = headers.get(header_key(rule.source))
                        raw_value = (
                            values[header[1]]
                            if header is not None and header[1] < len(values)
                            else rule.default
                        )
                        if rule.fill_down:
                            if _is_blank(raw_value):
                                raw_value = fill_down_values.get(
                                    rule.target,
                                    raw_value,
                                )
                            elif not _is_summary_label(raw_value):
                                fill_down_values[rule.target] = raw_value
                        mapped_row[rule.target] = _convert_value(raw_value, rule)
                    filter_row = {**raw_row, **mapped_row}
                    if self._row_passes_filters(filter_row):
                        output_rows.append(mapped_row)
                except ExtractionError as exc:
                    raise ExtractionError(
                        f"{path.name}, worksheet '{worksheet.title}', row {row_number}: {exc}"
                    ) from exc
            LOGGER.info("Extracted %d matching row(s) from %s", len(output_rows), path.name)
            return output_rows
        finally:
            workbook.close()

    def _select_worksheet(self, workbook: Any, path: Path) -> Any:
        sheet_name = self.config.input.sheet_name
        if sheet_name is None:
            return workbook.active
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        matches = [name for name in workbook.sheetnames if name.casefold() == sheet_name.casefold()]
        if len(matches) == 1:
            return workbook[matches[0]]
        raise ExtractionError(
            f"Workbook {path.name} has no worksheet named '{sheet_name}'. "
            f"Available: {', '.join(workbook.sheetnames)}"
        )

    def _row_passes_filters(self, row: Mapping[str, Any]) -> bool:
        if not self.config.filters.rules:
            return True
        results = [_matches_filter(row, rule) for rule in self.config.filters.rules]
        return all(results) if self.config.filters.mode == "all" else any(results)

    def _finalize(
        self,
        mapped_rows: list[dict[str, Any]],
        source_files: tuple[Path, ...],
    ) -> ExtractedData:
        if self.config.grouping.enabled:
            rows = self._group_rows(mapped_rows)
            headers = tuple(self.config.grouping.by) + tuple(
                aggregation.target for aggregation in self.config.grouping.aggregations
            )
        else:
            rows = list(mapped_rows)
            headers = tuple(
                rule.target for rule in self.config.columns if rule.include_in_report
            )

        if self.config.sort:
            rows.sort(key=cmp_to_key(self._compare_rows))

        formats, widths = self._output_metadata(headers)
        return ExtractedData(
            headers=headers,
            rows=tuple(rows),
            number_formats=formats,
            widths=widths,
            source_files=source_files,
        )

    def _group_rows(self, rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
        groups: OrderedDict[tuple[Any, ...], list[Mapping[str, Any]]] = OrderedDict()
        for row in rows:
            key = tuple(row.get(column) for column in self.config.grouping.by)
            groups.setdefault(key, []).append(row)

        output: list[dict[str, Any]] = []
        for key, group_rows in groups.items():
            result = dict(zip(self.config.grouping.by, key))
            for aggregation in self.config.grouping.aggregations:
                result[aggregation.target] = self._aggregate(group_rows, aggregation)
            output.append(result)
        return output

    @staticmethod
    def _aggregate(
        rows: Sequence[Mapping[str, Any]], aggregation: AggregationRule
    ) -> Any:
        if aggregation.operation == "count_rows":
            return len(rows)
        values = [
            row.get(aggregation.source) for row in rows if not _is_blank(row.get(aggregation.source))
        ]
        if aggregation.operation == "count":
            return len(values)
        if not values:
            return None
        if aggregation.operation in {"sum", "avg"}:
            if any(
                not isinstance(value, (int, float)) or isinstance(value, bool)
                for value in values
            ):
                raise ExtractionError(
                    f"Aggregation {aggregation.operation} requires numeric values in {aggregation.source}"
                )
            total = sum(values)
            return total if aggregation.operation == "sum" else total / len(values)
        if aggregation.operation == "min":
            return min(values)
        if aggregation.operation == "max":
            return max(values)
        raise ExtractionError(f"Unsupported aggregation: {aggregation.operation}")

    def _compare_rows(self, left: Mapping[str, Any], right: Mapping[str, Any]) -> int:
        for rule in self.config.sort:
            left_value = left.get(rule.column)
            right_value = right.get(rule.column)
            left_blank = _is_blank(left_value)
            right_blank = _is_blank(right_value)
            if left_blank and right_blank:
                continue
            if left_blank:
                return 1
            if right_blank:
                return -1
            if isinstance(left_value, str) or isinstance(right_value, str):
                left_comparable = str(left_value).casefold()
                right_comparable = str(right_value).casefold()
            else:
                left_comparable = left_value
                right_comparable = right_value
            if left_comparable < right_comparable:
                result = -1
            elif left_comparable > right_comparable:
                result = 1
            else:
                continue
            return -result if rule.direction == "desc" else result
        return 0

    def _output_metadata(
        self, headers: Sequence[str]
    ) -> tuple[dict[str, str], dict[str, float]]:
        columns_by_target = {rule.target: rule for rule in self.config.columns}
        aggregations_by_target = {
            rule.target: rule for rule in self.config.grouping.aggregations
        }
        formats: dict[str, str] = {}
        widths: dict[str, float] = {}
        for header in headers:
            column = columns_by_target.get(header)
            aggregation = aggregations_by_target.get(header)
            number_format = (
                aggregation.number_format
                if aggregation and aggregation.number_format
                else column.number_format if column else None
            )
            width = (
                aggregation.width
                if aggregation and aggregation.width is not None
                else column.width if column else None
            )
            if number_format:
                formats[header] = number_format
            if width is not None:
                widths[header] = width
        return formats, widths
