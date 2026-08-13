from __future__ import annotations

import logging
import os
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sop_reporter.config import ExtractionConfig, TextStyle
from sop_reporter.exceptions import ReportBuildError
from sop_reporter.extractor import ExtractedData


LOGGER = logging.getLogger(__name__)


def _font(style: TextStyle) -> Font:
    return Font(
        name=style.font_name,
        size=style.font_size,
        bold=style.bold,
        color=style.font_color,
    )


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _alignment(style: TextStyle) -> Alignment:
    return Alignment(horizontal=style.alignment, vertical="center", wrap_text=True)


def _body_alignment(style: TextStyle, value: object) -> Alignment:
    horizontal = style.alignment
    if isinstance(value, (date, datetime)):
        horizontal = "center"
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        horizontal = "right"
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=True)


class ReportBuilder:
    def __init__(self, config: ExtractionConfig) -> None:
        if config.report is None:
            raise ReportBuildError("Extraction configuration has no report section")
        self.config = config

    def title_for_partition(self, label: str) -> str:
        report = self.config.report
        assert report is not None
        if not self.config.split.enabled:
            return report.title
        return self.config.split.title_template.format(
            base_title=report.title,
            value=label,
        )

    def filename_suffix_for_partition(self, safe_label: str) -> str:
        report = self.config.report
        assert report is not None
        if not self.config.split.enabled:
            return ""
        return self.config.split.filename_suffix.format(
            base_title=report.title,
            value=safe_label,
        )

    def build(
        self,
        data: ExtractedData,
        destination: Path,
        *,
        generated_at: datetime | None = None,
        title: str | None = None,
    ) -> Path:
        if not data.headers:
            raise ReportBuildError("Cannot build a report without output columns")
        report = self.config.report
        assert report is not None
        generated_at = generated_at or datetime.now()
        effective_title = title or report.title
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)
        temporary = destination.with_name(
            f".{destination.stem}.{os.getpid()}.tmp.xlsx"
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = report.worksheet_name
        worksheet.sheet_view.showGridLines = False
        last_column = len(data.headers)
        last_column_letter = get_column_letter(last_column)

        try:
            self._write_title(
                worksheet,
                title=effective_title,
                last_column=last_column,
                style=report.title_style,
            )
            if report.show_generated_at:
                source_count = len(data.source_files)
                source_word = "source" if source_count == 1 else "sources"
                worksheet.cell(
                    row=2,
                    column=1,
                    value=(
                        f"{report.generated_at_label}: "
                        f"{generated_at:%m/%d/%Y %I:%M %p}  |  "
                        f"{source_count} {source_word}"
                    ),
                )
                if last_column > 1:
                    worksheet.merge_cells(
                        start_row=2,
                        start_column=1,
                        end_row=2,
                        end_column=last_column,
                    )
                worksheet.cell(row=2, column=1).font = Font(
                    name=report.body_style.font_name,
                    size=9,
                    italic=True,
                    color="5B6573",
                )

            header_row = report.table_start_row
            for column_index, header in enumerate(data.headers, start=1):
                cell = worksheet.cell(row=header_row, column=column_index, value=header)
                cell.font = _font(report.header_style)
                cell.fill = _fill(report.header_style.fill_color)
                cell.alignment = _alignment(report.header_style)
            if report.header_style.row_height is not None:
                worksheet.row_dimensions[header_row].height = report.header_style.row_height

            border = Border()
            if report.body_style.border_color:
                side = Side(style="thin", color=report.body_style.border_color)
                border = Border(left=side, right=side, top=side, bottom=side)

            for output_row, row in enumerate(data.rows, start=header_row + 1):
                is_alternate = (output_row - header_row) % 2 == 0
                row_fill_color = (
                    report.body_style.alternate_fill_color
                    if is_alternate and report.body_style.alternate_fill_color
                    else report.body_style.fill_color
                )
                for column_index, header in enumerate(data.headers, start=1):
                    cell = worksheet.cell(
                        row=output_row,
                        column=column_index,
                        value=row.get(header),
                    )
                    cell.font = _font(report.body_style)
                    cell.fill = _fill(row_fill_color)
                    cell.alignment = _body_alignment(report.body_style, cell.value)
                    cell.border = border
                    if header in data.number_formats:
                        cell.number_format = data.number_formats[header]
                if report.body_style.row_height is not None:
                    worksheet.row_dimensions[output_row].height = report.body_style.row_height

            for column_index, header in enumerate(data.headers, start=1):
                configured_width = data.widths.get(header)
                if configured_width is None:
                    values = [str(header)] + [
                        "" if row.get(header) is None else str(row.get(header))
                        for row in data.rows
                    ]
                    configured_width = min(45, max(10, max(map(len, values)) + 2))
                worksheet.column_dimensions[
                    get_column_letter(column_index)
                ].width = configured_width

            last_row = max(header_row, header_row + len(data.rows))
            table_reference = f"A{header_row}:{last_column_letter}{last_row}"
            if report.auto_filter:
                worksheet.auto_filter.ref = table_reference
            freeze_panes = report.freeze_panes
            worksheet.freeze_panes = (
                f"A{header_row + 1}" if freeze_panes.casefold() == "auto" else freeze_panes
            )

            page = report.page
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_TABLOID
            worksheet.page_setup.orientation = page.orientation
            worksheet.page_setup.fitToPage = True
            worksheet.page_setup.fitToWidth = page.fit_to_pages_wide
            worksheet.page_setup.fitToHeight = page.fit_to_pages_tall
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            worksheet.print_options.horizontalCentered = True
            worksheet.page_margins.left = page.margin_inches
            worksheet.page_margins.right = page.margin_inches
            worksheet.page_margins.top = page.margin_inches
            worksheet.page_margins.bottom = page.margin_inches
            worksheet.page_margins.header = 0.15
            worksheet.page_margins.footer = 0.15
            worksheet.print_title_rows = f"{header_row}:{header_row}"
            worksheet.print_area = f"A1:{last_column_letter}{last_row}"

            workbook.properties.title = effective_title
            workbook.properties.creator = "SOP Reporter"
            workbook.properties.created = generated_at
            workbook.save(temporary)
            os.replace(temporary, destination)
            LOGGER.info("Built report %s with %d row(s)", destination, len(data.rows))
            return destination
        except ReportBuildError:
            raise
        except Exception as exc:
            raise ReportBuildError(f"Could not build report {destination}: {exc}") from exc
        finally:
            workbook.close()
            try:
                temporary.unlink(missing_ok=True)
            except OSError:
                LOGGER.debug("Could not remove temporary report %s", temporary, exc_info=True)

    @staticmethod
    def _write_title(
        worksheet,
        *,
        title: str,
        last_column: int,
        style: TextStyle,
    ) -> None:
        worksheet.cell(row=1, column=1, value=title)
        if last_column > 1:
            worksheet.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=last_column,
            )
        for column_index in range(1, last_column + 1):
            cell = worksheet.cell(row=1, column=column_index)
            cell.fill = _fill(style.fill_color)
            cell.font = _font(style)
            cell.alignment = _alignment(style)
        worksheet.row_dimensions[1].height = style.row_height or 32
